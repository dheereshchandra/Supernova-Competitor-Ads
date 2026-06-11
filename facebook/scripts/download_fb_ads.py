#!/usr/bin/env python3
"""
Download videos from the Facebook Ad Library in bulk.

This is the workhorse for the `facebook-ad-downloader` skill. It takes a set of
Facebook ad references (library IDs, Ad Library page URLs, signed CDN URLs, a
Google Sheet, or a CSV/XLSX) and downloads every video into one folder with a
consistent naming scheme.

Why Ad Library page URLs are the reliable input
-----------------------------------------------
A Facebook ad creative lives on the CDN behind a *signed* URL with `oh` (an
HMAC signature) and `oe` (a hex Unix expiry) query params. The signature is
bound to the asset path and expires within hours, so a bare or stale CDN URL
will 403 and there is no way to forge it. The Ad Library *page* URL
(`facebook.com/ads/library/?id=<id>`) is stable: yt-dlp's `facebook:ads`
extractor re-resolves it to a fresh signed URL on every run. So whenever you
have a library ID, prefer building the page URL over chasing CDN links.

Input types (auto-detected per entry / per file)
-------------------------------------------------
  - A bare numeric library ID            -> built into an Ad Library page URL
  - facebook.com/ads/library/?id=...     -> resolved via yt-dlp
  - https://video.*.fbcdn.net/...mp4?oh= -> downloaded directly via HTTP
  - A .txt file                          -> one entry per line
  - A .csv / .tsv / .xlsx file           -> a column of entries
  - A public Google Sheets URL           -> exported to CSV, then a column

Output naming
-------------
One ad can have several creative variants under a single library ID. They are
saved as:
    <libraryId>.mp4      (primary)
    <libraryId>_v2.mp4   (second variant)
    <libraryId>_v3.mp4   ...

Resumability
------------
Re-running is safe: an ad whose `<libraryId>.mp4` already exists is skipped.
This also means a large job can be run in chunks with --limit / --offset, or
simply re-run after an interruption.

Usage
-----
    python download_fb_ads.py <source> [options]

    <source> is one of:
      - path to a .txt / .csv / .tsv / .xlsx file
      - a public Google Sheets URL
      - a single library ID or URL (handy for one-offs)

Options
-------
    --column NAME       Column holding the URL/ID (sheets & spreadsheets only).
                        Auto-detects ad_link / ad_url / *url* / *link* columns.
    --out DIR           Output directory (default: ./fb_ad_videos)
    --batch-size N      Library IDs per yt-dlp invocation (default: 10). yt-dlp
                        fetches a verification cookie once per call, so batching
                        amortizes that cost. Keep modest to stay clean of 429s.
    --limit N           Process only the first N entries (chunking / testing).
    --offset N          Skip the first N entries (chunking / resuming).
    --ytdlp PATH        Explicit path to the yt-dlp binary.
    --summary PATH      Where to write the run summary CSV
                        (default: <out>/download_summary.csv).

Exit code is 0 if every downloadable ad succeeded, 1 if any failed.
"""

import argparse
import csv
import io
import os
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from urllib.parse import urlparse, parse_qs


# --------------------------------------------------------------------------
# input loading
# --------------------------------------------------------------------------

def load_entries(source: str, column: str | None) -> list[str]:
    """Return a flat list of raw entry strings (IDs or URLs) from any source."""
    # Google Sheet
    if source.startswith("http") and "docs.google.com/spreadsheets" in source:
        m = re.search(r"/d/([a-zA-Z0-9-_]+)", source)
        if not m:
            sys.exit("Could not parse the Google Sheets ID out of that URL.")
        sheet_id = m.group(1)
        gid_m = re.search(r"[#&?]gid=(\d+)", source)
        gid = gid_m.group(1) if gid_m else "0"
        export = (f"https://docs.google.com/spreadsheets/d/{sheet_id}"
                  f"/export?format=csv&gid={gid}")
        print(f"[info] fetching Google Sheet as CSV: {export}")
        try:
            import requests
        except ImportError:
            sys.exit("Reading a Google Sheet needs: pip install requests")
        resp = requests.get(export, timeout=30)
        resp.raise_for_status()
        rows = list(csv.DictReader(io.StringIO(resp.text)))
        return _column_from_rows(rows, column)

    # A single inline ID or URL
    if source.startswith("http") or re.fullmatch(r"\d{8,}", source.strip()):
        return [source.strip()]

    path = Path(source)
    if not path.exists():
        sys.exit(f"Source not found: {source}")
    suffix = path.suffix.lower()

    if suffix in {".txt", ""}:
        return [ln.strip() for ln in path.read_text(encoding="utf-8-sig").splitlines()
                if ln.strip()]

    if suffix in {".csv", ".tsv"}:
        delim = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f, delimiter=delim))
        return _column_from_rows(rows, column)

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Reading .xlsx needs: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it)]
        rows = []
        for r in it:
            if all(c is None for c in r):
                continue
            rows.append({h: ("" if v is None else v) for h, v in zip(headers, r)})
        return _column_from_rows(rows, column)

    sys.exit(f"Unsupported file type: {suffix}")


def _column_from_rows(rows: list[dict], column: str | None) -> list[str]:
    if not rows:
        sys.exit("The sheet/spreadsheet has no data rows.")
    cols = list(rows[0].keys())
    if column:
        if column not in cols:
            sys.exit(f"Column {column!r} not found. Available columns: {cols}")
        chosen = column
    else:
        chosen = _detect_column(rows, cols)
        print(f"[info] auto-detected column: {chosen!r}")
    return [str(r.get(chosen, "")).strip() for r in rows if str(r.get(chosen, "")).strip()]


def _detect_column(rows: list[dict], cols: list[str]) -> str:
    # ad_link / page URLs are the most reliable, prefer them over raw media URLs.
    # ad_library_url is the post-2026-05-26 rename of ad_library_permalink (HANDOVER rename pass).
    for kw in ("ad_library_url", "ad_link", "ad_url", "library_url", "page_url"):
        for c in cols:
            if c and kw in c.lower():
                return c
    # ad_library_id is the post-rename name; libraryid / library_id / ad_id / id stay as fallbacks.
    for kw in ("adlibraryid", "ad_library_id", "libraryid", "library_id", "ad_id", "id"):
        for c in cols:
            if c and kw == c.lower().replace(" ", ""):
                return c
    for kw in ("url", "link", "video", "src"):
        for c in cols:
            if c and kw in c.lower():
                return c
    for c in cols:
        if str(rows[0].get(c, "")).startswith(("http://", "https://")):
            return c
    sys.exit(f"Could not auto-detect the URL/ID column. Pass --column. Columns: {cols}")


# --------------------------------------------------------------------------
# cleaning & normalisation
# --------------------------------------------------------------------------

SCI_RE = re.compile(r"^(\d+)\.(\d+)[eE]\+(\d+)$")


def clean_entry(raw: str) -> tuple[str | None, str]:
    """Normalise one raw entry.

    Returns (kind, value):
      ("id",  "<libraryId>")        a bare library ID, ready for an Ad Library URL
      ("page","<url>")              an Ad Library page URL
      ("cdn", "<url>")              a direct CDN URL
      ("sci", "<as-typed>")         Excel-mangled scientific notation - UNRESOLVED
      (None,  "<as-typed>")         unrecognised / unusable
    """
    s = raw.strip().strip('"').strip("'")
    if not s:
        return None, raw

    low = s.lower()
    if "facebook.com/ads/library" in low:
        return "page", s
    if "fbcdn.net" in low and ".mp4" in low:
        return "cdn", s

    # Excel turns 16-digit IDs into things like "1.97587E+15" - cannot be
    # recovered from the displayed form alone; the caller must map it back
    # against the source data.
    if SCI_RE.match(s):
        return "sci", s

    # Filenames / messy IDs: strip a .mp4(.part) extension, a trailing _N
    # version marker, and any non-digits.
    s = re.sub(r"\.mp4(\.part)?$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"_v?\d+$", "", s)          # _1, _2, _v2 version suffixes
    digits = re.sub(r"\D", "", s)
    if len(digits) >= 8:
        return "id", digits
    return None, raw


def ad_library_url(library_id: str) -> str:
    return f"https://www.facebook.com/ads/library/?id={library_id}"


# --------------------------------------------------------------------------
# CDN direct download
# --------------------------------------------------------------------------

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


def cdn_signed(url: str) -> tuple[bool, str]:
    qs = parse_qs(urlparse(url).query)
    missing = [p for p in ("oh", "oe") if p not in qs]
    if missing:
        return False, ("missing signature param(s): " + ", ".join(missing) +
                       " - bare CDN URLs always 403; use the Ad Library page URL")
    try:
        exp = int(qs["oe"][0], 16)
        if exp < time.time():
            return False, ("signature expired at " +
                            time.strftime("%Y-%m-%d %H:%M", time.gmtime(exp)) + " UTC")
    except ValueError:
        pass
    return True, "ok"


def download_cdn(url: str, dest: Path) -> tuple[bool, str]:
    if dest.exists() and dest.stat().st_size > 0:
        return True, f"skip (exists, {dest.stat().st_size:,} bytes)"
    ok, why = cdn_signed(url)
    if not ok:
        return False, why
    try:
        import requests
    except ImportError:
        return False, "needs: pip install requests"
    tmp = dest.with_suffix(dest.suffix + ".part")
    try:
        with requests.get(url, headers={"User-Agent": _UA, "Accept": "*/*"},
                          stream=True, timeout=60) as r:
            if r.status_code not in (200, 206):
                return False, f"HTTP {r.status_code}"
            total = 0
            with tmp.open("wb") as f:
                for chunk in r.iter_content(chunk_size=65536):
                    if chunk:
                        f.write(chunk)
                        total += len(chunk)
            tmp.rename(dest)
            return True, f"{total:,} bytes"
    except Exception as e:                       # noqa: BLE001
        # don't leave a half-written .part stub behind (best-effort)
        try:
            tmp.unlink()
        except OSError:
            pass
        return False, f"network error: {e.__class__.__name__}"


# --------------------------------------------------------------------------
# yt-dlp batch download for Ad Library pages
# --------------------------------------------------------------------------

def find_ytdlp(explicit: str | None) -> str:
    import shutil
    if explicit:
        return explicit
    found = shutil.which("yt-dlp")
    if found:
        return found
    for p in (Path.home() / ".local/bin/yt-dlp", Path("/usr/local/bin/yt-dlp")):
        if p.exists():
            return str(p)
    # yt-dlp installed as a module but not on PATH
    return sys.executable + " -m yt_dlp"


def run_ytdlp_batch(library_ids: list[str], out_dir: Path, ytdlp: str,
                    timeout: int = 600) -> tuple[set[str], dict[str, str]]:
    """Download a batch of Ad Library pages. Returns (succeeded_ids, {id: error})."""
    # The batch file goes in the system temp dir, NOT the output folder. Some
    # sandboxed or cloud-synced mounts block file *deletion*; writing the temp
    # file there and then deleting it would crash the run. tempfile lives on a
    # normal filesystem and is cleaned up in the finally block.
    fd, batch_path = tempfile.mkstemp(prefix="fb_ad_batch_", suffix=".txt")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as bf:
            bf.write("".join(ad_library_url(i) + "\n" for i in library_ids))
        # %(id)s for a facebook:ads playlist entry is "<libraryId>_<n>"; the
        # rename pass afterwards collapses that into <id>.mp4 / <id>_vN.mp4.
        out_tmpl = str(out_dir / "%(id)s.%(ext)s")
        cmd = (ytdlp.split() if " " in ytdlp else [ytdlp]) + [
            "--no-warnings", "--no-progress", "--ignore-errors",
            "--no-playlist-reverse",
            # Limit to first playlist entry per ad. Facebook's facebook:ads
            # extractor returns each aspect-ratio variant (9:16, 1:1, …) as a
            # separate playlist entry; for our master-CSV one-row-per-creative
            # model we only need the primary (first) entry. Multi-creative
            # ads (creative_count_in_ad > 1) are handled with a follow-up
            # pass that drops this constraint.
            "--playlist-items", "1",
            "--batch-file", batch_path,
            "-o", out_tmpl,
        ]
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True,
                                  timeout=timeout)
        except subprocess.TimeoutExpired:
            return set(), {i: "yt-dlp batch timed out" for i in library_ids}
    finally:
        try:
            os.unlink(batch_path)
        except OSError:
            pass

    errors: dict[str, str] = {}
    stderr = proc.stderr or ""
    # "ERROR: [facebook:ads] <id>: <message>"
    for m in re.finditer(r"\[facebook:ads\]\s+(\d+):\s*(.+)", stderr):
        lid, msg = m.group(1), m.group(2).strip()
        # only keep the first concise sentence of the message
        errors[lid] = msg.split("; please report")[0].split(". ")[0][:200]

    produced = {f.stem.split("_")[0] for f in out_dir.glob("*.mp4")}
    succeeded = {i for i in library_ids if i in produced}
    for i in library_ids:
        if i not in succeeded and i not in errors:
            errors[i] = "no file produced (unknown reason)"
    return succeeded, errors


# --------------------------------------------------------------------------
# rename pass: <id>_<n>.mp4  ->  <id>.mp4 / <id>_v2.mp4 ...
# --------------------------------------------------------------------------

def normalise_filenames(out_dir: Path) -> dict[str, int]:
    """Collapse yt-dlp's raw names into <id>.mp4 / <id>_vN.mp4. Returns {id: count}."""
    # Group every not-yet-normalised file by its leading library ID.
    groups: dict[str, list[Path]] = {}
    for f in out_dir.glob("*.mp4"):
        if re.fullmatch(r"\d+(_v\d+)?", f.stem):
            continue                      # already in <id>.mp4 / <id>_vN.mp4 form
        m = re.match(r"^(\d{8,})", f.stem)
        if m:
            groups.setdefault(m.group(1), []).append(f)

    for lid, files in groups.items():
        def order(p: Path) -> tuple:
            m = re.search(r"_p?(\d+)", p.stem[len(lid):])
            return (int(m.group(1)) if m else 0, p.name)
        files.sort(key=order)
        # de-dupe identical-size variants yt-dlp occasionally emits twice
        seen, uniq = set(), []
        for f in files:
            sz = f.stat().st_size
            if sz in seen:
                try:
                    f.unlink()
                    continue
                except OSError:
                    pass    # mount blocks deletion - keep it, give it a _vN slot
            seen.add(sz)
            uniq.append(f)
        # rename into the next free version slots, leaving any pre-existing
        # <id>*.mp4 files (from an earlier run) untouched
        existing = {p.name for p in out_dir.glob(f"{lid}*.mp4")
                    if re.fullmatch(rf"{lid}(_v\d+)?", p.stem)}
        idx = 1
        for f in uniq:
            while (f"{lid}.mp4" if idx == 1 else f"{lid}_v{idx}.mp4") in existing:
                idx += 1
            target = f"{lid}.mp4" if idx == 1 else f"{lid}_v{idx}.mp4"
            if f.name != target:
                f.rename(out_dir / target)
            existing.add(target)
            idx += 1

    # Count fresh, from the final state of the folder - one pass, no accumulation.
    counts: dict[str, int] = {}
    for f in out_dir.glob("*.mp4"):
        m = re.match(r"^(\d{8,})", f.stem)
        if m:
            counts[m.group(1)] = counts.get(m.group(1), 0) + 1
    return counts


def cleanup_partials(out_dir: Path) -> int:
    """Best-effort sweep of yt-dlp leftovers: *.part / *.ytdl files and 0-byte
    *.mp4 stubs. Returns how many were removed. Deletion failures (read-only or
    no-delete mounts) are swallowed - this is housekeeping, never fatal."""
    removed = 0
    leftovers = list(out_dir.glob("*.part")) + list(out_dir.glob("*.ytdl"))
    for f in out_dir.glob("*.mp4"):
        try:
            if f.stat().st_size == 0:
                leftovers.append(f)
        except OSError:
            pass
    for f in leftovers:
        try:
            f.unlink()
            removed += 1
        except OSError:
            pass
    return removed


# --------------------------------------------------------------------------
# version-aware pass (v2.3 CSVs): one media file per (version, creative)
# --------------------------------------------------------------------------

def _int0(x) -> int:
    try:
        return int(str(x or "0").strip() or "0")
    except (ValueError, TypeError):
        return 0


def media_name(ad_id: str, v, c, ext: str = ".mp4") -> str:
    """Version-aware media filename — MUST match upload_to_r2.media_name /
    probe_media / transcribe_tag: version 0 keeps the legacy {id}.mp4 (creative 0) /
    {id}_v{N+1}.mp4 (creative N); version V>=1 -> {id}_ver{V}_c{C}.mp4."""
    vi, ci = _int0(v), _int0(c)
    if vi <= 0:
        return f"{ad_id}{ext}" if ci == 0 else f"{ad_id}_v{ci + 1}{ext}"
    return f"{ad_id}_ver{vi}_c{ci}{ext}"


_V23_COLS = {"ad_library_id", "version_index", "creative_index_in_ad",
             "facebook_video_cdn_url_at_scrape"}


def _file_sha(p: Path) -> str | None:
    try:
        import hashlib
        return hashlib.sha256(p.read_bytes()).hexdigest()
    except OSError:
        return None


def download_versions_pass(csv_path: str, out_dir: Path, results: dict) -> tuple[int, int, int, int]:
    """For a v2.3 input CSV, fetch every (version, creative) video straight from its
    captured `facebook_video_cdn_url_at_scrape` into the version-aware filename. yt-dlp
    only fetches an ad's DEFAULT version (it can't drive FB's in-detail version nav), so
    this fills in extra versions ({id}_ver{V}_c{C}.mp4) and is a CDN fallback for any
    default the yt-dlp pass missed.

    DEDUP: FB frequently enumerates 2+ "versions" of an ad that are the SAME underlying
    video (different version_id / signed URL, identical bytes — verified across a whole
    SpeakX scrape). To avoid duplicate media in R2 (and duplicate transcription), an extra
    version whose URL matches the default, or whose downloaded bytes are identical to the
    default, is dropped — the version row keeps its metadata but shares the default media.
    Only GENUINELY distinct version creatives are kept as separate files.

    Returns (downloaded, skipped_present, failed, deduped). No-op unless v2.3 columns."""
    try:
        with open(csv_path, encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.DictReader(fh))
    except (OSError, csv.Error):
        return 0, 0, 0, 0
    if not rows or not _V23_COLS.issubset(rows[0].keys()):
        return 0, 0, 0, 0

    by_ad: dict[str, list] = {}
    for r in rows:
        if "video" not in (r.get("ad_media_type", "") or "").strip().lower():
            continue
        aid = (r.get("ad_library_id") or "").strip()
        url = (r.get("facebook_video_cdn_url_at_scrape") or "").strip()
        if aid and url.startswith("http"):
            by_ad.setdefault(aid, []).append(r)

    got = skip = fail = dedup = 0
    for aid, rs in by_ad.items():
        default = out_dir / f"{aid}.mp4"
        drow = next((r for r in rs if media_name(aid, r.get("version_index"),
                     r.get("creative_index_in_ad")) == f"{aid}.mp4"), None)
        if (not default.exists() or default.stat().st_size == 0) and drow:
            ok, detail = download_cdn(drow["facebook_video_cdn_url_at_scrape"].strip(), default)
            results[default.name] = ("downloaded" if ok else "failed", detail)
            got += 1 if ok else 0
            fail += 0 if ok else 1
        dsha = _file_sha(default) if default.exists() else None
        seen_urls = {drow["facebook_video_cdn_url_at_scrape"].strip()} if drow else set()
        for r in rs:
            name = media_name(aid, r.get("version_index"), r.get("creative_index_in_ad"))
            if name == f"{aid}.mp4":
                continue
            dest = out_dir / name
            if dest.exists() and dest.stat().st_size > 0:
                skip += 1
                continue
            url = r["facebook_video_cdn_url_at_scrape"].strip()
            if url in seen_urls:                 # same signed URL as default → same media
                dedup += 1
                continue
            seen_urls.add(url)
            ok, detail = download_cdn(url, dest)
            if not ok:
                results[name] = ("failed", detail)
                fail += 1
                continue
            if dsha and _file_sha(dest) == dsha:  # identical bytes to default → drop the dup
                dest.unlink(missing_ok=True)
                dedup += 1
                continue
            results[name] = ("downloaded", detail)  # genuinely distinct version creative
            got += 1
    return got, skip, fail, dedup


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("source", help="ID/URL, .txt/.csv/.xlsx file, or Google Sheets URL")
    ap.add_argument("--column")
    ap.add_argument("--out", default="./fb_ad_videos")
    ap.add_argument("--batch-size", type=int, default=10)
    ap.add_argument("--limit", type=int)
    ap.add_argument("--offset", type=int, default=0)
    ap.add_argument("--ytdlp")
    ap.add_argument("--summary")
    args = ap.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    raw = load_entries(args.source, args.column)
    if args.offset:
        raw = raw[args.offset:]
    if args.limit:
        raw = raw[:args.limit]

    pages_by_id: dict[str, str] = {}     # libraryId -> page url (deduped)
    cdn: list[str] = []
    sci: list[str] = []
    bad: list[str] = []
    for entry in raw:
        kind, val = clean_entry(entry)
        if kind == "id":
            pages_by_id.setdefault(val, ad_library_url(val))
        elif kind == "page":
            m = re.search(r"[?&]id=(\d+)", val)
            pages_by_id.setdefault(m.group(1) if m else val, val)
        elif kind == "cdn":
            cdn.append(val)
        elif kind == "sci":
            sci.append(val)
        else:
            bad.append(val)

    print(f"[info] {len(pages_by_id)} Ad Library ad(s), {len(cdn)} direct CDN URL(s)")
    if sci:
        print(f"[warn] {len(sci)} entries are Excel scientific-notation and CANNOT be")
        print( "       used as-is. Recover the full IDs from the source data first:")
        for s in sci:
            print(f"         {s}")
    if bad:
        print(f"[warn] {len(bad)} unrecognised entr(ies) skipped: {bad[:5]}"
              + (" ..." if len(bad) > 5 else ""))

    ytdlp = find_ytdlp(args.ytdlp)
    print(f"[info] yt-dlp: {ytdlp}")
    print(f"[info] output: {out_dir.resolve()}\n")

    results: dict[str, tuple[str, str]] = {}    # id/url -> (status, detail)

    # ----- Ad Library pages, in batches -----
    todo = [i for i in pages_by_id
            if not (out_dir / f"{i}.mp4").exists()]
    skipped = [i for i in pages_by_id if i not in todo]
    for i in skipped:
        results[i] = ("skipped", "already downloaded")
    if skipped:
        print(f"[info] {len(skipped)} ad(s) already downloaded - skipping")

    for start in range(0, len(todo), args.batch_size):
        batch = todo[start:start + args.batch_size]
        n = start // args.batch_size + 1
        total = (len(todo) + args.batch_size - 1) // args.batch_size
        print(f"[batch {n}/{total}] {len(batch)} ad(s)...")
        ok, errs = run_ytdlp_batch(batch, out_dir, ytdlp)
        for i in batch:
            if i in ok:
                results[i] = ("downloaded", "")
            else:
                results[i] = ("failed", errs.get(i, "unknown error"))
        print(f"             {len(ok)} ok, {len(batch) - len(ok)} failed")

    # Retry pass. The facebook:ads extractor is flaky - a fresh request a few
    # seconds later sometimes succeeds where the first one returned "Unable to
    # extract ad data". One extra pass is cheap insurance; anything still
    # failing after it is very likely a genuinely removed ad.
    retry = sorted(i for i, (s, _) in results.items() if s == "failed")
    if retry:
        print(f"\n[retry] re-attempting {len(retry)} failed ad(s) after a short wait...")
        time.sleep(8)
        for start in range(0, len(retry), args.batch_size):
            batch = retry[start:start + args.batch_size]
            ok, errs = run_ytdlp_batch(batch, out_dir, ytdlp)
            for i in batch:
                if i in ok:
                    results[i] = ("downloaded", "(recovered on retry)")
            if ok:
                print(f"        recovered {len(ok)}: {sorted(ok)}")

    cleanup_partials(out_dir)
    counts = normalise_filenames(out_dir)

    # ----- direct CDN URLs -----
    for idx, url in enumerate(cdn, 1):
        name = re.sub(r"[^\w.-]+", "_", Path(urlparse(url).path).stem) or f"cdn_{idx}"
        dest = out_dir / f"{name}.mp4"
        ok, detail = download_cdn(url, dest)
        results[url] = ("downloaded" if ok else "failed", detail)
        print(f"[cdn {idx}/{len(cdn)}] {'ok' if ok else 'FAIL'} - {detail}")

    # ----- version-aware pass (v2.3 CSVs): extra versions/creatives via captured CDN -----
    if args.source.lower().endswith(".csv"):
        vgot, vskip, vfail, vdedup = download_versions_pass(args.source, out_dir, results)
        if vgot or vfail or vskip or vdedup:
            print(f"[versions] {vgot} distinct version/creative file(s) downloaded, "
                  f"{vdedup} identical version(s) deduped, {vskip} already present, "
                  f"{vfail} failed (expired/blocked CDN)")

    # ----- summary -----
    summary_path = Path(args.summary) if args.summary else out_dir / "download_summary.csv"
    failed_path = out_dir / "failed_ads.txt"
    with summary_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["entry", "status", "versions", "detail"])
        for key, (status, detail) in sorted(results.items()):
            w.writerow([key, status, counts.get(key, "" if status == "failed" else 1),
                        detail])
    failures = sorted(k for k, (s, _) in results.items() if s == "failed")
    if failures:
        failed_path.write_text("\n".join(failures) + "\n")

    n_ok = sum(1 for s, _ in results.values() if s in ("downloaded", "skipped"))
    n_fail = len(failures)
    n_files = len(list(out_dir.glob("*.mp4")))
    print(f"\n[done] {n_ok} ad(s) available, {n_fail} failed, {n_files} video file(s)")
    print(f"[done] folder:  {out_dir.resolve()}")
    print(f"[done] summary: {summary_path.resolve()}")
    if failures:
        print(f"[done] failed:  {failed_path.resolve()}")
        print('[note] "Unable to extract ad data" usually means the ad is no longer')
        print("       published on Facebook (advertiser deleted it / it was removed).")
        print("       The facebook:ads extractor is also occasionally flaky, so a few")
        print("       may recover on a later run - re-run the script and it will retry")
        print("       only what is still missing (already-downloaded ads are skipped).")
    # exit code: PARTIAL failures are NON-FATAL — expired/removed-ad CDN links are
    # expected, the succeeded ads are in hand, and the master carries forward the
    # rest. Halting the whole competitor on a few dead links was the old all-or-
    # nothing trap. Exit non-zero ONLY if the download wholly failed (nothing came
    # down while ads were attempted) — a real outage worth stopping for.
    if n_fail:
        print(f"[warn] {n_fail} download(s) failed — surfaced, NON-FATAL ({n_ok} available).")
    sys.exit(1 if (n_ok == 0 and n_fail > 0) else 0)


if __name__ == "__main__":
    main()
