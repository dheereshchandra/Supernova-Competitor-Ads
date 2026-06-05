#!/usr/bin/env python3
"""
Scrape and download Google Ads Transparency Center creatives in bulk.

Takes one or more Google advertiser IDs (the AR... strings) and produces, for
each advertiser, a CSV of every creative + a folder of downloaded videos and
images. Replaces the Claude-in-Chrome scraper step entirely — talks to the
Transparency Center's public RPC endpoints directly, no browser automation.

Input formats (auto-detected)
-----------------------------
  - inline:   one or more AR... advertiser IDs as positional args
  - file:     path to a .txt file (one AR... per line)
  - file:     path to a .csv / .xlsx with an `advertiser_id` column
  - sheet:    public Google Sheets URL with an `advertiser_id` column
  - shortcut: `--competitor <name>` to use the built-in competitor map (the
              same 9 verified IDs from the Supernova competitor list)

What gets pulled per advertiser
-------------------------------
For every creative the API returns (paged in batches of 40, all pages):
  - format    YouTubeVideo / Image / Text / Other
  - last shown / first shown dates
  - regions the ad ran in
  - per-variant content URL (Google's content.js preview link)
  - video URL extracted from VAST (signed googlevideo.com mp4) — for videos
  - YouTube video ID + watch URL — derived by base64url-encoding the 8-byte
    hex `id` parameter from the googlevideo URL (confirmed reversible)
  - image URL — for image ads (tpc.googlesyndication.com/archive/simgad/...)

Then it downloads:
  - YouTubeVideo: googlevideo signed mp4 → videos/{slug}-{date}/{cr_id}.mp4
                  (variants saved as {cr_id}_v2.mp4, _v3.mp4, ...)
  - Image:        simgad image → images/{slug}-{date}/{cr_id}.{jpg,png,gif}
  - Text/Other:   no asset, captured in the CSV only

CSV output
----------
Writes inputs/g-ads-{slug}-{date}.csv with this header (chosen to match the
existing Supernova Step 1 schema, so downstream Step 2 / Step 3 pipelines
work unchanged):

  row_rank, competitor_name, advertiser_id, advertiser_name,
  advertiser_verified_location, target_country, creative_id, creative_url,
  creative_format, creative_last_shown_date, regions_shown, variant_index,
  variant_count, ad_headline, ad_description, ad_cta_text, ad_overlay_text,
  ad_destination_url, youtube_video_id, youtube_video_url,
  youtube_video_title, youtube_video_description, youtube_channel_name,
  youtube_video_duration_s, image_url_at_scrape, aspect_ratio,
  scrape_run_date, error_tag

Resumability
------------
Re-runs are safe. The script checks the output folder before each download
and skips files already on disk. So `--limit` / interruptions / a second run
will only pick up what's missing.

Usage
-----
  # one advertiser by ID:
  python3 scrape_google_ads.py AR02221466555617640449 --slug speakx

  # multiple advertisers, India-only:
  python3 scrape_google_ads.py AR012... AR042... --slug duolingo --region IN

  # competitor shortcut (uses the verified-AR map below):
  python3 scrape_google_ads.py --competitor SpeakX

  # all 9 verified competitors in one batch:
  python3 scrape_google_ads.py --all-competitors

  # from a file:
  python3 scrape_google_ads.py advertiser_ids.txt --slug mybatch
"""
from __future__ import annotations

import argparse
import base64
import csv
import datetime
import io
import json
import os
import pathlib
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
import urllib.request
from typing import Optional

import random

# Local module — rate-limit guardrail + run ledger (see scrape_guardrail.py).
# Adding the script's own dir to sys.path makes this import work regardless of cwd.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import scrape_guardrail as guardrail  # noqa: E402

# Third-party (auto-installed if missing) — see ensure_deps()
try:
    import requests  # noqa: F401
except ImportError:
    requests = None  # type: ignore


# ---------------------------------------------------------------------------
# Competitor → verified AR... advertiser ID map.
# Mirrors the Supernova find_advertiser_ids discovery output; edit here when
# the team confirms more IDs (e.g. the second Duolingo entity, Loora ×3).
# ---------------------------------------------------------------------------

COMPETITOR_ADVERTISERS: dict[str, list[dict]] = {
    "MySivi": [
        # ⚠ name mismatch — verify with MySivi team
        {"advertiser_id": "AR00309923070552834049",
         "advertiser_name": "NinjaSalary Financial Services Private Limited",
         "location": "India"},
    ],
    "SpeakX": [
        # ⚠ name mismatch — brand verifies under Ivypods
        {"advertiser_id": "AR02221466555617640449",
         "advertiser_name": "Ivypods Technology Pvt. Ltd",
         "location": "India"},
    ],
    "English Seekho": [
        # ⚠ name mismatch — brand verifies under Keyaro
        {"advertiser_id": "AR02289573295139323905",
         "advertiser_name": "Keyaro Edutech Private Limited",
         "location": "India"},
    ],
    "Speak": [
        {"advertiser_id": "AR00705977088842137601",
         "advertiser_name": "Speakeasy Labs, Inc",
         "location": "United States"},
    ],
    "Loora": [
        {"advertiser_id": "AR13725632235724341249",
         "advertiser_name": "Loora AI LTD",
         "location": "Israel"},
    ],
    "Duolingo": [
        {"advertiser_id": "AR01320432650854334465",
         "advertiser_name": "Duolingo Inc",
         "location": "United States"},
    ],
    "Busuu": [
        {"advertiser_id": "AR10122694483049971713",
         "advertiser_name": "Busuu Limited",
         "location": "United Kingdom"},
    ],
    "Memrise": [
        {"advertiser_id": "AR06420837225457516545",
         "advertiser_name": "Memrise Ltd",
         "location": "United Kingdom"},
    ],
    "Praktika AI": [
        {"advertiser_id": "AR10065453821808607233",
         "advertiser_name": "PRAKTIKA.AI COMPANY",
         "location": "United States"},
    ],
}

# Region code map — the Transparency Center API takes a numeric Geographic
# Targeting code per the internal proto. The full list is in regions.py of
# the public reverse-engineered library; we only need the ones we use.
# `None` → no region filter (the API returns ads from any region).
REGION_CODES: dict[str, int] = {
    "IN": 2356, "US": 2840, "GB": 2826, "CA": 2124, "AU": 2036,
    "DE": 2276, "FR": 2250, "JP": 2392, "BR": 2076, "ES": 2724,
    "IT": 2380, "MX": 2484, "ID": 2360, "PH": 2608, "AE": 2784,
    "SG": 2702, "PK": 2586, "BD": 2050, "LK": 2144, "NP": 2524,
}

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


# ---------------------------------------------------------------------------
# Dependency bootstrapping — keep the script runnable as a single file
# ---------------------------------------------------------------------------

def ensure_deps() -> None:
    """Install missing third-party deps. Idempotent / no-op when present.

    Tries pip variants in order so the script works across Apple CLT
    Python 3.9 (no --break-system-packages), Homebrew Python 3.11+ (needs
    --break-system-packages or --user to bypass PEP 668), and virtualenvs.
    """
    global requests  # noqa: PLW0603
    missing = []
    try:
        import requests as _r  # noqa: F401
        requests = _r  # type: ignore
    except ImportError:
        missing.append("requests")
    if not missing:
        return
    print(f"[setup] installing missing deps: {missing}")
    attempts = [
        ["--user", "--quiet"],
        ["--break-system-packages", "--quiet"],
        ["--quiet"],
    ]
    last_err = None
    for extra in attempts:
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", *extra, *missing],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            last_err = None
            break
        except subprocess.CalledProcessError as e:
            last_err = e
            continue
    if last_err is not None:
        sys.exit(
            f"[setup] could not install {missing}. "
            f"Run manually: pip3 install --user {' '.join(missing)}")
    # On --user installs the import path may need a refresh
    import site, importlib  # noqa: E401
    importlib.reload(site)
    import requests as _r  # noqa: F401, F811
    requests = _r  # type: ignore


# ---------------------------------------------------------------------------
# Transparency Center API client
# ---------------------------------------------------------------------------

class TransparencyClient:
    """Talks to adstransparency.google.com's internal SearchService /
    LookupService RPCs. Mirrors the request shape used by the SPA — the same
    shape the open-source Google-Ads-Transparency-Scraper PyPI module uses.

    Region behaviour: the API actually applies the region filter server-side
    (despite some reverse-engineering write-ups claiming otherwise). Passing
    region_code=None gives the "Anywhere" view; an integer code (e.g. 2356
    for India) restricts results to that region.
    """

    BASE = "https://adstransparency.google.com"
    SEARCH_URL = BASE + "/anji/_/rpc/SearchService/SearchCreatives"
    DETAIL_URL = BASE + "/anji/_/rpc/LookupService/GetCreativeById"

    def __init__(self, region_code: Optional[int]):
        self.region_code = region_code
        self.session = requests.Session()  # type: ignore[attr-defined]
        self.session.headers.update({
            "user-agent": UA,
            "accept": "*/*",
            "accept-language": "en-US,en;q=0.9",
            "origin": self.BASE,
            "x-same-domain": "1",
        })
        # Warm cookies — Google sets a NID / GTM cookie on the first GET; without
        # this the POST sometimes returns an empty {} body.
        self.session.get(self.BASE + "/", params={"region": "anywhere"}, timeout=20)

    def search_creatives(self, advertiser_id: str, page_size: int = 40,
                         next_page_id: Optional[str] = None) -> dict:
        """One page of creatives for an advertiser. Returns the parsed JSON."""
        body: dict = {
            "2": min(page_size, 100),
            "3": {
                "12": {"1": "", "2": True},
                "13": {"1": [advertiser_id]},
            },
            "7": {"1": 1},
        }
        if self.region_code is not None:
            body["3"]["8"] = [self.region_code]
        if next_page_id:
            body["4"] = next_page_id
        return self._post(self.SEARCH_URL, body)

    def list_all_creatives(self, advertiser_id: str,
                           max_count: Optional[int] = None) -> list[dict]:
        """All creatives for an advertiser, paging until exhausted."""
        results: list[dict] = []
        next_page: Optional[str] = None
        while True:
            resp = self.search_creatives(advertiser_id, next_page_id=next_page)
            page = resp.get("1") or []
            results.extend(page)
            next_page = resp.get("2")
            if not next_page or not page:
                break
            if max_count and len(results) >= max_count:
                break
            time.sleep(guardrail.SEARCH_PAGE_DELAY_SECONDS
                       + random.uniform(0, guardrail.SEARCH_PAGE_JITTER))  # paced + jittered
        if max_count:
            results = results[:max_count]
        return results

    def get_creative_detail(self, advertiser_id: str, creative_id: str) -> dict:
        """Full detail for one creative: format, all variants, regions, etc."""
        body = {"1": advertiser_id, "2": creative_id, "5": {"1": 1}}
        return self._post(self.DETAIL_URL, body)

    def _post(self, url: str, body: dict) -> dict:
        """POST with automatic 429 backoff. Google's anti-abuse occasionally
        redirects to /sorry/index with 429 mid-run; instead of crashing, we
        sleep an increasing amount and retry. Total wait ceiling is ~7 min,
        after which we raise — by then the IP is probably block-listed and
        the user needs to back off entirely."""
        data = {"f.req": json.dumps(body, separators=(",", ":"))}
        backoffs = [15, 30, 60, 120, 240]  # seconds — cumulative ~7 min
        for attempt, wait in enumerate([0] + backoffs):
            if wait:
                print(f"          [429] backing off {wait}s "
                      f"(attempt {attempt}/{len(backoffs)})", flush=True)
                time.sleep(wait)
            r = self.session.post(url, params={"authuser": "0"},
                                  data=data, timeout=30,
                                  allow_redirects=False)
            # 302 to /sorry/index is Google's captcha block — same family as 429
            if r.status_code in (429, 302, 503):
                ra = r.headers.get("Retry-After", "")
                if ra.isdigit():
                    extra = min(int(ra), 600)
                    print(f"          [429] server asked Retry-After={extra}s; "
                          f"honoring", flush=True)
                    time.sleep(extra)
                continue
            r.raise_for_status()
            try:
                return r.json()
            except json.JSONDecodeError:
                return {}
        # Out of attempts
        raise requests.exceptions.HTTPError(  # type: ignore[attr-defined]
            f"429 persisted after {sum(backoffs)}s of backoff — IP likely "
            f"block-listed; wait 30+ min before retrying. URL: {url}")


# ---------------------------------------------------------------------------
# Creative parsing — extract the actual video / image / text from API payload
# ---------------------------------------------------------------------------

# The Transparency Center's `format` integer (detail["8"]) does NOT cleanly map
# to (Text, Image, Video). In practice format_int=1 and =2 both appear for image
# banner ads, and =3 means HTML5/video. The reliable approach is to inspect
# the variant payload itself: if any variant has a content.js URL whose body
# contains a googlevideo.com link → YouTubeVideo; if any has an <img src=...>
# HTML snippet → Image; otherwise → Text / Other.
FORMAT_HINT = {1: "Image", 2: "Image", 3: "YouTubeVideo"}


def parse_creative_detail(detail: dict) -> dict:
    """Take the GetCreativeById response and pull out everything we need.

    Returns a dict with:
      format          "YouTubeVideo" / "Image" / "Text" / "Other"
      last_shown      "YYYY-MM-DD" or ""
      first_shown     "YYYY-MM-DD" or ""
      regions         list of region codes (numeric)
      advertiser_name string
      variants        list of per-variant dicts (one per variation Google has):
                        { content_url, image_url, ad_html }
    """
    inner = detail.get("1") or {}
    fmt_int = inner.get("8")
    # Initial hint from format_int — refined below once we see what the
    # variants actually contain (the int alone is unreliable, see comment above).
    fmt = FORMAT_HINT.get(fmt_int, "Text")

    # Last shown — top-level field "4" is the last-shown timestamp
    last_shown = _ts_to_date(inner.get("4"))
    # Regions — "17" is a list of {"1": region_code, "5": YYYYMMDD last shown}
    regions = []
    for entry in inner.get("17") or []:
        rc = entry.get("1")
        if rc is not None:
            regions.append(rc)
    # First shown can sometimes be inferred from "17" entries' "4" field
    first_shown = ""
    for entry in inner.get("17") or []:
        first = entry.get("4")
        if isinstance(first, int) and first > 0:
            first_shown = _yyyymmdd_to_date(first)
            break

    # Advertiser name — "22.1"
    advertiser_name = (inner.get("22") or {}).get("1") or ""

    variants = []
    for v in inner.get("5") or []:
        content_url = ""
        image_url = ""
        ad_html = ""
        # Video / display formats: content URL at v["1"]["4"]
        if "1" in v and isinstance(v["1"], dict):
            content_url = v["1"].get("4") or ""
        # Image format: HTML snippet at v["3"]["2"] (an <img src=...> string)
        if "3" in v and isinstance(v["3"], dict):
            ad_html = v["3"].get("2") or ""
            m = re.search(r'src=["\'](https?://[^"\']+)', ad_html)
            if m:
                image_url = m.group(1)
        variants.append({
            "content_url": content_url,
            "image_url": image_url,
            "ad_html": ad_html,
        })

    # Refine the format based on what the variants actually carry. An <img src=>
    # snippet → Image regardless of format_int; a content.js URL with no image
    # snippet → likely YouTubeVideo (confirmed once we fetch the body and find a
    # googlevideo URL — see the main loop). Text is the catch-all.
    has_image = any(v["image_url"] for v in variants)
    has_content_url = any(v["content_url"] for v in variants)
    if has_image:
        fmt = "Image"
    elif has_content_url and fmt_int == 3:
        fmt = "YouTubeVideo"
    elif has_content_url:
        # Display/HTML5 with no obvious image src — treat as Image candidate;
        # we'll downgrade to Text if the content.js turns out to be text-only.
        fmt = "Image"

    return {
        "format": fmt,
        "format_int": fmt_int,
        "last_shown": last_shown,
        "first_shown": first_shown,
        "regions": regions,
        "advertiser_name": advertiser_name,
        "variants": variants,
    }


def _ts_to_date(field: Optional[dict]) -> str:
    """The API uses {"1": "<unix_secs_string>", "2": <nanos>} for timestamps."""
    if not isinstance(field, dict):
        return ""
    secs = field.get("1")
    try:
        return datetime.datetime.fromtimestamp(int(secs)).strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        return ""


def _yyyymmdd_to_date(n: int) -> str:
    s = str(n)
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return ""


# ---------------------------------------------------------------------------
# Video URL extraction — content.js → VAST → googlevideo mp4 → YouTube ID
# ---------------------------------------------------------------------------

# Per-run content.js cache. Many variants of one creative share the same
# preview URL (only assets= and htmlParentId= differ); caching saves dozens
# of redundant 100 KB HTTP fetches on a multi-variant advertiser.
_CONTENT_JS_CACHE: dict[str, str] = {}


def fetch_content_js(content_url: str, session) -> str:
    """Fetch a content.js preview URL; decode \\xNN escape sequences once.

    Cached per-URL for the duration of the script run. Returns the decoded
    text body (large — typically 80-200 KB). Returns empty string on any
    network error so a single bad URL doesn't tank the whole scrape.
    """
    # Strip the htmlParentId/responseCallback query params — they only affect
    # the page DOM the JS would render into, not the ad content, but they
    # make otherwise-identical content URLs cache-miss.
    cache_key = re.sub(r"&(?:htmlParentId|responseCallback)=[^&]*", "",
                       content_url)
    if cache_key in _CONTENT_JS_CACHE:
        return _CONTENT_JS_CACHE[cache_key]
    try:
        r = session.get(content_url, timeout=10,
                        headers={"user-agent": UA, "accept": "*/*"})
        if r.status_code != 200:
            _CONTENT_JS_CACHE[cache_key] = ""
            return ""
        body = re.sub(
            r"\\x([0-9a-fA-F]{2})", lambda m: chr(int(m.group(1), 16)),
            r.text)
    except Exception:  # noqa: BLE001
        body = ""
    _CONTENT_JS_CACHE[cache_key] = body
    return body


def extract_googlevideo_url(content_text: str) -> str:
    """Pull the signed googlevideo.com/videoplayback URL out of the VAST body.

    Returns "" if none found. The URL is short-lived (~5h) so callers should
    download immediately, not store the URL for later.
    """
    m = re.search(r'https?://[^\s"<>\\]*googlevideo\.com/videoplayback[^\s"<>\\\]]+', content_text)
    return m.group(0) if m else ""


def youtube_id_from_googlevideo(url: str) -> str:
    """The `id=<16-hex>` parameter in a googlevideo URL is the YouTube video ID
    encoded as 8 bytes of hex. base64url-encoding those 8 bytes yields the
    standard 11-char YouTube watch ID.

    Verified: id=98fc7471cc4ff5a8 → mPx0ccxP9ag (matches existing data).
    Returns "" if the URL has no id param or the param isn't 16 hex chars.
    """
    m = re.search(r"[?&]id=([0-9a-fA-F]{16})(?:&|$)", url)
    if not m:
        return ""
    try:
        b = bytes.fromhex(m.group(1))
        return base64.urlsafe_b64encode(b).decode("ascii").rstrip("=")
    except ValueError:
        return ""


# Google chrome / housekeeping hosts that are never an ad's real destination.
# Used to filter out telemetry / policy / preview-tool links from clickthrough
# candidates.
_GOOGLE_CHROME_HOSTS = (
    "google.com/about", "google.com/intl", "google.com/policies",
    "policies.google.com", "support.google.com", "ads.google.com",
    "googleadservices.com", "doubleclick.net", "googlesyndication.com",
    "googleusercontent.com", "google-analytics.com", "googletagmanager.com",
    "gstatic.com", "safety.google", "myaccount.google.com",
    "myadcenter.google.com", "blog.google", "adservice.google.",
    "adstransparency.google.com", "youtube.com", "youtu.be", "ggpht.com",
    # Asset / font CDNs Google ads pull in but that are never the destination
    "fonts.googleapis.com", "fonts.gstatic.com", "ajax.googleapis.com",
    "maps.googleapis.com", "google.com/recaptcha",
)


def _is_chrome_url(url: str) -> bool:
    low = url.lower()
    return any(host in low for host in _GOOGLE_CHROME_HOSTS)


def extract_clickthrough_url(content_text: str) -> str:
    """Pull the ad's destination (clickthrough) URL out of a content.js body.

    Works across all three Transparency Center content types:
      - VAST video: `<ClickThrough><![CDATA[...]]></ClickThrough>`
      - Display/HTML5: a top-level `<a href="...">` wrapping the creative
      - Text ad: same — the rendered HTML has an outbound anchor

    We strip Google-chrome / tracking hosts so the first surviving candidate
    is the real advertiser destination.
    """
    # VAST ClickThrough is the most authoritative when present
    m = re.search(
        r"<ClickThrough[^>]*>\s*(?:<!\[CDATA\[)?(https?://[^\s<\]]+)",
        content_text, re.IGNORECASE)
    if m and not _is_chrome_url(m.group(1)):
        return m.group(1)
    # Fallback: any outbound href that isn't Google chrome
    for h in re.finditer(r'href=["\'](https?://[^"\']+)["\']', content_text):
        url = h.group(1)
        if not _is_chrome_url(url):
            return url
    # Last-ditch: any AdServingFinalUrl / clickThroughUrl JS field
    m = re.search(
        r'(?:clickThroughUrl|adServingFinalUrl)["\']?\s*[:=]\s*["\']'
        r'(https?://[^"\']+)', content_text, re.IGNORECASE)
    if m and not _is_chrome_url(m.group(1)):
        return m.group(1)
    return ""


def extract_text_ad_copy(content_text: str) -> dict:
    """Best-effort headline / description / destination for ANY format ad.

    The content.js for text/search ads contains escaped HTML with the rendered
    ad. For display/HTML5 ads, headline-like text is rarer (the creative is
    usually pre-rendered as an image) but we still try. For video ads the
    destination comes from the VAST ClickThrough element.

    Returns {"headline": "", "description": "", "destination": ""}.
    """
    out = {"headline": "", "description": "", "destination": ""}
    out["destination"] = extract_clickthrough_url(content_text)

    # First reasonably long div / span / heading with visible text → headline.
    # Skip chrome strings (Material icon names, common SPA labels). Also reject
    # anything that's all-caps and >50 chars (usually a fingerprint / token),
    # generic UI labels ("Downloads", "Install", etc.), and tracker IDs.
    _UI_LABELS = {
        "video", "image", "ad", "ads", "advertisement", "downloads",
        "install", "open", "play", "share", "subscribe", "watch", "view",
        "more", "back", "next", "close", "menu", "search", "submit",
    }
    for tag_m in re.finditer(
            r'<(?:div|span|h[1-4]|p)[^>]*>([^<]{8,160})</', content_text):
        txt = tag_m.group(1).strip()
        if not txt:
            continue
        if txt.startswith("&") or txt.lower().startswith("http"):
            continue
        if txt.lower() in _UI_LABELS:
            continue
        if txt.isupper() and len(txt) > 50:
            continue
        # Tracker IDs and base64-ish tokens
        if re.fullmatch(r"[A-Za-z0-9_+/=-]{20,}", txt):
            continue
        if not out["headline"]:
            out["headline"] = txt
        elif not out["description"] and txt != out["headline"]:
            out["description"] = txt
            break
    return out


# ---------------------------------------------------------------------------
# Downloaders — direct HTTP (no yt-dlp needed)
# ---------------------------------------------------------------------------

def download_file(url: str, dest: pathlib.Path, session,
                  referer: str = "https://adstransparency.google.com/") -> tuple[bool, str]:
    """Stream a URL to dest. Returns (ok, detail).

    Idempotent: if dest already exists and is >10KB, skip immediately."""
    if dest.exists() and dest.stat().st_size > 10_000:
        return True, f"skipped (already {dest.stat().st_size:,} bytes)"
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(dest.suffix + ".part")
    try:
        with session.get(url, stream=True, timeout=120,
                         headers={"user-agent": UA, "accept": "*/*",
                                  "referer": referer}) as r:
            if r.status_code not in (200, 206):
                return False, f"HTTP {r.status_code}"
            total = 0
            with tmp.open("wb") as f:
                for chunk in r.iter_content(chunk_size=65536):
                    if chunk:
                        f.write(chunk)
                        total += len(chunk)
        if total < 1000:
            tmp.unlink()
            return False, f"too small ({total} bytes) — probably a placeholder"
        tmp.rename(dest)
        return True, f"{total:,} bytes"
    except Exception as e:  # noqa: BLE001
        try:
            tmp.unlink()
        except OSError:
            pass
        return False, f"network error: {e.__class__.__name__}: {e}"


def guess_image_ext(url: str, content_type: str = "") -> str:
    """Pick a sensible .jpg / .png / .gif / .webp from the URL or MIME type."""
    low = url.lower()
    for ext in (".png", ".gif", ".webp", ".jpeg", ".jpg"):
        if ext in low:
            return ".jpg" if ext == ".jpeg" else ext
    if "png" in content_type:
        return ".png"
    if "gif" in content_type:
        return ".gif"
    if "webp" in content_type:
        return ".webp"
    return ".jpg"


# ---------------------------------------------------------------------------
# Input loading — IDs from positional args / files / sheets
# ---------------------------------------------------------------------------

def load_advertiser_ids(source: str) -> list[dict]:
    """Return a list of {advertiser_id, advertiser_name?, location?} dicts.

    `source` is one of: a raw AR... ID, a path to .txt/.csv/.xlsx, a public
    Google Sheets URL."""
    s = source.strip()
    if re.fullmatch(r"AR[0-9A-Za-z]{15,}", s):
        return [{"advertiser_id": s}]

    if s.startswith("http") and "docs.google.com/spreadsheets" in s:
        m = re.search(r"/d/([a-zA-Z0-9-_]+)", s)
        if not m:
            sys.exit(f"Cannot parse Google Sheet ID from {s}")
        gid_m = re.search(r"[#&?]gid=(\d+)", s)
        gid = gid_m.group(1) if gid_m else "0"
        export = (f"https://docs.google.com/spreadsheets/d/{m.group(1)}"
                  f"/export?format=csv&gid={gid}")
        resp = requests.get(export, timeout=30)  # type: ignore[attr-defined]
        resp.raise_for_status()
        rows = list(csv.DictReader(io.StringIO(resp.text)))
        return _rows_to_advertisers(rows)

    p = pathlib.Path(s)
    if not p.exists():
        sys.exit(f"Source not found: {s}")
    suf = p.suffix.lower()

    if suf in (".txt", ""):
        ids: list[dict] = []
        for line in p.read_text(encoding="utf-8-sig").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            m = re.search(r"AR[0-9A-Za-z]{15,}", line)
            if m:
                ids.append({"advertiser_id": m.group(0)})
        return ids

    if suf in (".csv", ".tsv"):
        delim = "\t" if suf == ".tsv" else ","
        with p.open(encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f, delimiter=delim))
        return _rows_to_advertisers(rows)

    if suf in (".xlsx", ".xlsm", ".xls"):
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip", "install",
                                   "--break-system-packages", "--quiet",
                                   "openpyxl"])
            from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it)]
        rows = []
        for r in it:
            if all(c is None for c in r):
                continue
            rows.append({h: ("" if v is None else v) for h, v in zip(headers, r)})
        return _rows_to_advertisers(rows)

    sys.exit(f"Unsupported file type: {suf}")


def _rows_to_advertisers(rows: list[dict]) -> list[dict]:
    """Pick advertiser_id (+ optional name/location) columns out of any sheet."""
    if not rows:
        return []
    cols = list(rows[0].keys())

    def pick(*keywords) -> Optional[str]:
        for c in cols:
            cn = str(c).lower().replace(" ", "_")
            if any(k in cn for k in keywords):
                return c
        return None

    id_col = pick("advertiser_id", "ar_id", "ariid", "advertiser id")
    if not id_col:
        # Fall back to scanning any column for an AR... pattern
        for c in cols:
            if rows and re.search(r"AR[0-9A-Za-z]{15,}",
                                  str(rows[0].get(c) or "")):
                id_col = c
                break
    if not id_col:
        sys.exit(f"Could not find advertiser_id column. Columns: {cols}")
    name_col = pick("advertiser_name", "name", "verified_name")
    loc_col = pick("location", "country", "region")

    out: list[dict] = []
    seen: set[str] = set()
    for r in rows:
        raw = str(r.get(id_col) or "").strip()
        m = re.search(r"AR[0-9A-Za-z]{15,}", raw)
        if not m or m.group(0) in seen:
            continue
        seen.add(m.group(0))
        out.append({
            "advertiser_id": m.group(0),
            "advertiser_name": str(r.get(name_col) or "").strip() if name_col else "",
            "location": str(r.get(loc_col) or "").strip() if loc_col else "",
        })
    return out


# ---------------------------------------------------------------------------
# Main scrape loop
# ---------------------------------------------------------------------------

CSV_HEADER = [
    "row_rank", "competitor_name", "advertiser_id", "advertiser_name",
    "advertiser_verified_location", "target_country", "creative_id",
    "creative_url", "creative_format", "creative_last_shown_date",
    "regions_shown", "variant_index", "variant_count", "ad_headline",
    "ad_description", "ad_cta_text", "ad_overlay_text", "ad_destination_url",
    "youtube_video_id", "youtube_video_url", "youtube_video_title",
    "youtube_video_description", "youtube_channel_name",
    "youtube_video_duration_s", "image_url_at_scrape", "aspect_ratio",
    "scrape_run_date", "error_tag",
]


def scrape(args: argparse.Namespace) -> int:
    today = datetime.date.today().isoformat()
    slug = re.sub(r"[^a-z0-9]+", "-", (args.slug or "batch").lower()).strip("-")
    out_root = pathlib.Path(args.out or ".").resolve()
    inputs_dir = out_root / "inputs"
    videos_dir = out_root / "videos" / f"{slug}-{today}"
    images_dir = out_root / "images" / f"{slug}-{today}"
    inputs_dir.mkdir(parents=True, exist_ok=True)

    region_code = None
    if args.region:
        region_code = REGION_CODES.get(args.region.upper())
        if region_code is None:
            sys.exit(f"Unknown region {args.region!r}. Known: "
                     f"{sorted(REGION_CODES.keys())}")

    advertisers = _resolve_advertisers(args)
    if not advertisers:
        sys.exit("No advertiser IDs to scrape.")

    print(f"[info] competitor: {args.competitor or slug}")
    print(f"[info] advertisers: {len(advertisers)} "
          f"({', '.join(a['advertiser_id'] for a in advertisers)})")
    print(f"[info] region: {args.region or 'Anywhere'}")
    print(f"[info] output: {out_root}")
    print()

    # ── Rate-limit guardrail ────────────────────────────────────────────────
    # Enforce the cadence operating policy BEFORE any network call, and record
    # this run to logs/scrape_history.jsonl so the policy survives across
    # sessions. See scrape_guardrail.py and HANDOVER "Rate-limit operating
    # policy". Bypass with --force (no prompts) or --no-guardrail (skip).
    if not guardrail.enforce(out_root, competitor_count=len(advertisers),
                             force=getattr(args, "force", False),
                             no_guardrail=getattr(args, "no_guardrail", False)):
        return 3

    client = TransparencyClient(region_code=region_code)

    rows: list[dict] = []
    total_creatives = 0
    video_jobs: list[tuple[str, str, str, str, int]] = []  # (advid, crid, video_url, watch_id, variant_idx)
    image_jobs: list[tuple[str, str, str, int]] = []        # (advid, crid, image_url, variant_idx)

    for adv in advertisers:
        adv_id = adv["advertiser_id"]
        print(f"[search] {adv_id} ({adv.get('advertiser_name') or '?'})")
        try:
            creatives = client.list_all_creatives(adv_id, max_count=args.limit)
        except requests.exceptions.HTTPError as e:  # 429 persisted → IP block
            guardrail.record_run(
                out_root, competitor=slug,
                advertisers=[a["advertiser_id"] for a in advertisers],
                region=args.region, outcome="blocked",
                creatives=total_creatives, rows=len(rows), note=str(e)[:200])
            print(f"\n[BLOCKED] {e}")
            print(f"[guardrail] 429 block recorded — a "
                  f"{guardrail.POST_BLOCK_COOLDOWN_HOURS:.0f}h cooldown is now "
                  f"in effect (see: python3 scripts/scrape_guardrail.py status).")
            return 3
        total_creatives += len(creatives)
        print(f"          → {len(creatives)} creatives")
        for i, c in enumerate(creatives, 1):
            cr_id = c.get("2")
            if not cr_id:
                continue
            try:
                detail = client.get_creative_detail(adv_id, cr_id)
            except Exception as e:  # noqa: BLE001
                print(f"          ! detail failed for {cr_id}: {e}")
                rows.append(_error_row(adv, cr_id, args, today,
                                       f"detail-fetch: {e.__class__.__name__}"))
                continue
            parsed = parse_creative_detail(detail)
            adv_name = (adv.get("advertiser_name")
                        or parsed.get("advertiser_name") or "")
            variants = parsed["variants"] or [{}]
            n_variants = len(variants)
            for v_idx, variant in enumerate(variants):
                content_url = variant.get("content_url") or ""
                image_url = variant.get("image_url") or ""
                yt_id = ""
                video_url = ""
                ad_copy = {"headline": "", "description": "", "destination": ""}
                aspect = ""

                # Any variant with a content.js URL gets fetched once. The body
                # is parsed for: VAST video URL (video ads only), clickthrough
                # destination, and rendered headline/description text. This
                # used to be branched per-format but most banner / HTML5 ads
                # also carry a ClickThrough URL that should land in
                # ad_destination_url — so we run the extractor universally.
                if content_url:
                    body = fetch_content_js(content_url, client.session)
                    if parsed["format"] == "YouTubeVideo":
                        gv = extract_googlevideo_url(body)
                        if gv:
                            video_url = gv
                            yt_id = youtube_id_from_googlevideo(gv)
                        am = re.search(
                            r'\b(\d{1,2}):(\d{1,2})\b(?=.{0,200}video)', body)
                        if am:
                            aspect = f"{am.group(1)}:{am.group(2)}"
                    ad_copy = extract_text_ad_copy(body)

                # Queue downloads
                if video_url:
                    video_jobs.append((adv_id, cr_id, video_url, yt_id, v_idx))
                if image_url and parsed["format"] == "Image":
                    image_jobs.append((adv_id, cr_id, image_url, v_idx))

                # Diagnostic error_tag — explains why r2_public_url will be empty
                # downstream. Empty for rows whose asset will be uploaded. Step 3
                # leaves these alone (it only overwrites on its own successful
                # upload path). See HANDOVER.md §10.
                row_error_tag = ""
                fmt_here = parsed["format"]
                if fmt_here == "YouTubeVideo":
                    if not video_url:
                        row_error_tag = "html5-banner-no-mp4"
                elif fmt_here == "Image":
                    if not image_url:
                        row_error_tag = "image-url-not-extracted"
                elif fmt_here == "Text":
                    row_error_tag = "text-no-asset"
                elif fmt_here in ("", "Unknown", "Other"):
                    row_error_tag = "unknown-no-format"

                rows.append({
                    "row_rank": 0,  # filled after sort
                    "competitor_name": args.competitor or slug,
                    "advertiser_id": adv_id,
                    "advertiser_name": adv_name,
                    "advertiser_verified_location": adv.get("location") or "",
                    "target_country": args.region or "",
                    "creative_id": cr_id,
                    "creative_url": (
                        f"https://adstransparency.google.com/advertiser/"
                        f"{adv_id}/creative/{cr_id}"
                        + (f"?region={args.region}" if args.region else "")),
                    "creative_format": parsed["format"],
                    "creative_last_shown_date": parsed["last_shown"],
                    "regions_shown": ",".join(str(r) for r in parsed["regions"]),
                    "variant_index": v_idx,
                    "variant_count": n_variants,
                    "ad_headline": ad_copy["headline"],
                    "ad_description": ad_copy["description"],
                    "ad_cta_text": "",
                    "ad_overlay_text": "",
                    "ad_destination_url": ad_copy["destination"],
                    "youtube_video_id": yt_id,
                    "youtube_video_url": (f"https://www.youtube.com/watch?v={yt_id}"
                                          if yt_id else ""),
                    "youtube_video_title": "",
                    "youtube_video_description": "",
                    "youtube_channel_name": "",
                    "youtube_video_duration_s": "",
                    "image_url_at_scrape": image_url,
                    "aspect_ratio": aspect,
                    "scrape_run_date": today,
                    "error_tag": row_error_tag,
                })
            if i % 20 == 0:
                print(f"          processed {i}/{len(creatives)}", flush=True)
            time.sleep(guardrail.DETAIL_DELAY_SECONDS
                       + random.uniform(0, guardrail.DETAIL_JITTER))

    # Sort + rank rows
    rows.sort(key=lambda r: (r["advertiser_id"], r["creative_id"],
                             r["variant_index"]))
    for i, r in enumerate(rows, 1):
        r["row_rank"] = i

    csv_path = inputs_dir / f"g-ads-{slug}-{today}.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_HEADER, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"\n[csv] wrote {len(rows)} rows → {csv_path}")

    if args.no_download:
        print("[info] --no-download set; skipping asset downloads")
        guardrail.record_run(
            out_root, competitor=slug,
            advertisers=[a["advertiser_id"] for a in advertisers],
            region=args.region, outcome="success",
            creatives=total_creatives, rows=len(rows),
            note="no-download (CSV only)")
        return 0

    n_video_ok = n_video_fail = 0
    if video_jobs:
        print(f"\n[downloads] {len(video_jobs)} video creative(s) to fetch")
        videos_dir.mkdir(parents=True, exist_ok=True)
        for j, (adv_id, cr_id, url, yt_id, v_idx) in enumerate(video_jobs, 1):
            suffix = "" if v_idx == 0 else f"_v{v_idx + 1}"
            dest = videos_dir / f"{cr_id}{suffix}.mp4"
            ok, detail = download_file(url, dest, client.session)
            print(f"  [{j}/{len(video_jobs)}] {cr_id}{suffix}.mp4  "
                  f"{'OK ' if ok else 'FAIL'} — {detail}")
            if ok:
                n_video_ok += 1
            else:
                n_video_fail += 1
            time.sleep(0.2)

    n_image_ok = n_image_fail = 0
    if image_jobs:
        print(f"\n[downloads] {len(image_jobs)} image creative(s) to fetch")
        images_dir.mkdir(parents=True, exist_ok=True)
        for j, (adv_id, cr_id, url, v_idx) in enumerate(image_jobs, 1):
            suffix = "" if v_idx == 0 else f"_v{v_idx + 1}"
            ext = guess_image_ext(url)
            dest = images_dir / f"{cr_id}{suffix}{ext}"
            ok, detail = download_file(url, dest, client.session)
            print(f"  [{j}/{len(image_jobs)}] {cr_id}{suffix}{ext}  "
                  f"{'OK ' if ok else 'FAIL'} — {detail}")
            if ok:
                n_image_ok += 1
            else:
                n_image_fail += 1
            time.sleep(0.15)

    print()
    print(f"[done] CSV:    {csv_path}")
    if video_jobs:
        print(f"[done] videos: {videos_dir} — {n_video_ok} ok, {n_video_fail} failed")
    if image_jobs:
        print(f"[done] images: {images_dir} — {n_image_ok} ok, {n_image_fail} failed")
    print(f"[done] rows:   {len(rows)} (text/other variants stored in CSV only)")
    guardrail.record_run(
        out_root, competitor=slug,
        advertisers=[a["advertiser_id"] for a in advertisers],
        region=args.region, outcome="success",
        creatives=total_creatives, rows=len(rows),
        note=f"videos {n_video_ok}ok/{n_video_fail}fail, "
             f"images {n_image_ok}ok/{n_image_fail}fail")
    return 1 if (n_video_fail + n_image_fail) else 0


def _error_row(adv: dict, cr_id: str, args: argparse.Namespace, today: str,
               err: str) -> dict:
    return {
        "row_rank": 0,
        "competitor_name": args.competitor or args.slug or "",
        "advertiser_id": adv["advertiser_id"],
        "advertiser_name": adv.get("advertiser_name") or "",
        "advertiser_verified_location": adv.get("location") or "",
        "target_country": args.region or "",
        "creative_id": cr_id,
        "creative_url": (
            f"https://adstransparency.google.com/advertiser/"
            f"{adv['advertiser_id']}/creative/{cr_id}"),
        "creative_format": "Unknown",
        "creative_last_shown_date": "",
        "regions_shown": "",
        "variant_index": 0,
        "variant_count": 1,
        "ad_headline": "",
        "ad_description": "",
        "ad_cta_text": "",
        "ad_overlay_text": "",
        "ad_destination_url": "",
        "youtube_video_id": "",
        "youtube_video_url": "",
        "youtube_video_title": "",
        "youtube_video_description": "",
        "youtube_channel_name": "",
        "youtube_video_duration_s": "",
        "image_url_at_scrape": "",
        "aspect_ratio": "",
        "scrape_run_date": today,
        "error_tag": err,
    }


def _resolve_advertisers(args: argparse.Namespace) -> list[dict]:
    """Combine all input sources (positional args, --competitor, --all-competitors)."""
    out: list[dict] = []
    seen: set[str] = set()

    if args.all_competitors:
        for name, entries in COMPETITOR_ADVERTISERS.items():
            for e in entries:
                if e["advertiser_id"] not in seen:
                    seen.add(e["advertiser_id"])
                    out.append(dict(e, _competitor=name))
        return out

    if args.competitor:
        entries = COMPETITOR_ADVERTISERS.get(args.competitor)
        if not entries:
            sys.exit(
                f"Unknown competitor {args.competitor!r}. Known: "
                f"{sorted(COMPETITOR_ADVERTISERS.keys())}")
        for e in entries:
            if e["advertiser_id"] not in seen:
                seen.add(e["advertiser_id"])
                out.append(dict(e))

    for src in args.sources:
        for e in load_advertiser_ids(src):
            if e["advertiser_id"] not in seen:
                seen.add(e["advertiser_id"])
                out.append(e)

    return out


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("sources", nargs="*",
                    help="Advertiser IDs, files (.txt/.csv/.xlsx), or Google Sheets URLs")
    ap.add_argument("--competitor",
                    help="Use the built-in competitor map (e.g. SpeakX, Duolingo)")
    ap.add_argument("--all-competitors", action="store_true",
                    help="Process every entry in the competitor map")
    ap.add_argument("--slug",
                    help="Slug for the output CSV / folder name (default: based on competitor or 'batch')")
    ap.add_argument("--region",
                    help="Region code to restrict ads to (e.g. IN, US). Default = Anywhere.")
    ap.add_argument("--out", default=".",
                    help="Output root (default: current dir). Creates inputs/, videos/{slug-date}/, images/{slug-date}/")
    ap.add_argument("--limit", type=int,
                    help="Cap creatives per advertiser (for quick tests)")
    ap.add_argument("--no-download", action="store_true",
                    help="Skip the asset downloads (CSV only)")
    ap.add_argument("--force", action="store_true",
                    help="Override the rate-limit guardrail without prompts "
                         "(use only if you accept the IP-block risk)")
    ap.add_argument("--no-guardrail", action="store_true",
                    help="Skip the rate-limit guardrail entirely (testing only)")
    args = ap.parse_args()

    if not args.sources and not args.competitor and not args.all_competitors:
        ap.error("Provide at least one source (advertiser ID / file) "
                 "or --competitor / --all-competitors")
    if not args.slug and args.competitor:
        args.slug = re.sub(r"[^a-z0-9]+", "-", args.competitor.lower()).strip("-")
    if not args.slug and args.all_competitors:
        args.slug = "all-competitors"

    ensure_deps()
    return scrape(args)


if __name__ == "__main__":
    sys.exit(main())
